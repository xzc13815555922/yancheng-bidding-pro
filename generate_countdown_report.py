#!/usr/bin/env python3
"""
generate_countdown_report.py — 盐开开标倒计时报告

数据来源：unified.db tender 表
过滤：std_district IN ('盐南','经开') + proj_major_cat IS NULL + open_date >= 当月1日
输出：output/盐开开标倒计时报告_YYYYMMDD.xlsx
"""

import sqlite3
from datetime import date, datetime
from pathlib import Path
from typing import Optional

try:
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side
    )
    from openpyxl.utils import get_column_letter
except ImportError:
    raise SystemExit("缺少 openpyxl：pip install openpyxl")

DATA_DIR   = Path(__file__).parent / "data"
OUTPUT_DIR = Path(__file__).parent / "output"
UNIFIED_DB = DATA_DIR / "unified.db"

# ── 样式常量 ────────────────────────────────────────────────
HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT  = Font(name="微软雅黑", bold=True, color="FFFFFF", size=10)
TITLE_FONT   = Font(name="微软雅黑", bold=True, size=13, color="1F4E79")
CELL_FONT    = Font(name="微软雅黑", size=10)
LINK_FONT    = Font(name="微软雅黑", size=10, color="0563C1", underline="single")
PAST_FILL    = PatternFill("solid", fgColor="F2F2F2")   # 已过期（开标时间已过）
SOON_FILL    = PatternFill("solid", fgColor="FFF2CC")   # 3天内开标
ALT_FILL     = PatternFill("solid", fgColor="EBF3FB")   # 隔行淡蓝
URGENT_FONT  = Font(name="微软雅黑", size=10, bold=True, color="C00000")
BORDER_THIN  = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

# 列定义：(列头, 宽度, 对齐)
COLUMNS = [
    ("网站",         22, "left"),
    ("项目名称",     48, "left"),
    ("发布时间",     13, "center"),
    ("发包人",       28, "left"),
    ("预算金额(万元)", 14, "center"),
    ("开标时间",     18, "center"),
    ("倒计时(天)",   11, "center"),
    ("项目链接",     18, "center"),
]


def _query(today_str: str) -> list[dict]:
    month_start = today_str[:7] + "-01"
    conn = sqlite3.connect(str(UNIFIED_DB))
    conn.row_factory = sqlite3.Row
    rows = conn.execute(
        """
        SELECT site_name, project_name, publish_date,
               purchaser, budget, open_date, detail_url
        FROM tender
        WHERE proj_major_cat IS NULL
          AND proj_minor_cat IS NULL
          AND std_district IN ('盐南', '经开')
          AND open_date >= ?
        ORDER BY open_date ASC, publish_date ASC
        """,
        (month_start,),
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def _countdown(open_date_str: str, today: date) -> Optional[int]:
    """返回距开标天数（已过为负数）。"""
    if not open_date_str:
        return None
    try:
        dt = datetime.strptime(open_date_str[:10], "%Y-%m-%d").date()
        return (dt - today).days
    except ValueError:
        return None


def _fmt_budget(budget) -> str:
    if budget is None:
        return ""
    wan = budget / 10000
    return f"{wan:.2f}" if wan != int(wan) else f"{int(wan)}"


def _fmt_open(open_date_str: str) -> str:
    if not open_date_str:
        return ""
    try:
        dt = datetime.strptime(open_date_str, "%Y-%m-%d %H:%M:%S")
        return dt.strftime("%m-%d %H:%M")
    except ValueError:
        return open_date_str[:16]


def build(output_path: Optional[Path] = None):
    today      = date.today()
    today_str  = today.strftime("%Y-%m-%d")
    rows       = _query(today_str)

    OUTPUT_DIR.mkdir(exist_ok=True)
    if output_path is None:
        output_path = OUTPUT_DIR / f"盐开开标倒计时报告_{today.strftime('%Y%m%d')}.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "开标倒计时"

    # ── 标题行 ──────────────────────────────────────────────
    ws.merge_cells(f"A1:{get_column_letter(len(COLUMNS))}1")
    title_cell = ws["A1"]
    title_cell.value = f"盐开 · 开标倒计时清单（未分类项目）  {today_str}"
    title_cell.font      = TITLE_FONT
    title_cell.alignment = ALIGN_CENTER
    ws.row_dimensions[1].height = 28

    # 副标题：数量统计
    ws.merge_cells(f"A2:{get_column_letter(len(COLUMNS))}2")
    past    = sum(1 for r in rows if _countdown(r["open_date"], today) is not None and _countdown(r["open_date"], today) < 0)
    future  = len(rows) - past
    sub = ws["A2"]
    sub.value = f"共 {len(rows)} 条  |  即将开标 {future} 条  |  已过期 {past} 条（含当月历史）"
    sub.font      = Font(name="微软雅黑", size=9, italic=True, color="595959")
    sub.alignment = ALIGN_CENTER
    ws.row_dimensions[2].height = 18

    # ── 表头行 ──────────────────────────────────────────────
    for col_idx, (header, width, _) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = ALIGN_CENTER
        cell.border    = BORDER_THIN
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[3].height = 22
    ws.freeze_panes = "A4"

    # ── 数据行 ──────────────────────────────────────────────
    for row_idx, r in enumerate(rows, start=4):
        cd = _countdown(r["open_date"], today)

        # 行底色
        if cd is None or cd < 0:
            row_fill = PAST_FILL
        elif cd <= 3:
            row_fill = SOON_FILL
        elif row_idx % 2 == 0:
            row_fill = ALT_FILL
        else:
            row_fill = None

        values = [
            r["site_name"]    or "",
            r["project_name"] or "",
            (r["publish_date"] or "")[:10],
            r["purchaser"]    or "",
            _fmt_budget(r["budget"]),
            _fmt_open(r["open_date"]),
            "" if cd is None else cd,
            "",  # 链接列单独处理
        ]
        aligns = [c[2] for c in COLUMNS]

        for col_idx, (val, align) in enumerate(zip(values, aligns), 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border    = BORDER_THIN
            cell.alignment = ALIGN_CENTER if align == "center" else ALIGN_LEFT

            # 倒计时列特殊字体
            if col_idx == 7 and isinstance(val, int):
                if val < 0:
                    cell.font = Font(name="微软雅黑", size=10, color="999999")
                elif val <= 3:
                    cell.font = URGENT_FONT
                else:
                    cell.font = CELL_FONT
            else:
                cell.font = CELL_FONT

            if row_fill:
                cell.fill = row_fill

        # 链接列（超链接）
        url = r.get("detail_url") or ""
        link_cell = ws.cell(row=row_idx, column=8)
        if url:
            link_cell.hyperlink = url
            link_cell.value     = "查看公告"
            link_cell.font      = LINK_FONT
        else:
            link_cell.value = ""
            link_cell.font  = CELL_FONT
        link_cell.alignment = ALIGN_CENTER
        link_cell.border    = BORDER_THIN
        if row_fill:
            link_cell.fill = row_fill

        ws.row_dimensions[row_idx].height = 32

    # ── 图例说明 ─────────────────────────────────────────────
    legend_row = len(rows) + 5
    ws.merge_cells(f"A{legend_row}:D{legend_row}")
    legend = ws[f"A{legend_row}"]
    legend.value     = "■ 黄色：距开标 ≤3天    ■ 灰色：已过开标时间    ■ 蓝色：正常待开标"
    legend.font      = Font(name="微软雅黑", size=9, color="595959", italic=True)
    legend.alignment = ALIGN_LEFT

    wb.save(output_path)
    print(f"✓ 已生成：{output_path}  ({len(rows)} 条记录)")
    return output_path


if __name__ == "__main__":
    build()
