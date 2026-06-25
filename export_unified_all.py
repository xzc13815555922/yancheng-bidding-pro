#!/usr/bin/env python3
"""导出 unified.db 三张子表（tender / award / intention）全列全数据 Excel

azE 要求：
- 数据源：unified.db
- 3 张表各自 1 个 Sheet + 1 个汇总 Sheet
- 全列全行不脱敏
- 报告 open_date 字段填充率
"""
import sqlite3
from pathlib import Path
from datetime import datetime
import sys

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

DATA_DIR = Path(__file__).parent / "data"
OUTPUT_DIR = Path(__file__).parent / "output"

TABLES = [
    ("tender",    "招标公告"),
    ("award",     "中标/成交公告"),
    ("intention", "采购意向"),
]


def fetch_table(table_name: str):
    db_path = DATA_DIR / "unified.db"
    db = sqlite3.connect(str(db_path))
    try:
        cols_info = db.execute(f"PRAGMA table_info({table_name})").fetchall()
        cols = [r[1] for r in cols_info]
        if not cols:
            db.close()
            return None, None, None
        rows = db.execute(f"SELECT * FROM {table_name}").fetchall()
        count = db.execute(f"SELECT COUNT(*) FROM {table_name}").fetchone()[0]
        # open_date 填充率（仅 tender 表有 open_date 字段）
        has_open_date = "open_date" in cols
        open_date_filled = 0
        open_date_filled_pct = 0.0
        if has_open_date:
            idx = cols.index("open_date")
            open_date_filled = sum(
                1 for r in rows if r[idx] is not None and str(r[idx]).strip() != ""
            )
            open_date_filled_pct = round(open_date_filled * 100.0 / count, 2) if count else 0.0
        db.close()
        df = pd.DataFrame(rows, columns=cols)
        return df, count, {
            "cols": cols,
            "has_open_date": has_open_date,
            "open_date_filled": open_date_filled,
            "open_date_filled_pct": open_date_filled_pct,
        }
    except Exception as e:
        print(f"  ✗ {table_name} 失败: {e}", file=sys.stderr)
        db.close()
        return None, None, None


def main():
    print("=" * 60)
    print("unified.db 三表全量导出（azE 任务 v2）")
    print("=" * 60)

    OUTPUT_DIR.mkdir(exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    out_path = OUTPUT_DIR / f"统一招标公告数据库_{ts}.xlsx"

    summary_rows = []
    table_results = []

    for table_key, table_label in TABLES:
        print(f"  → {table_label} ({table_key})")
        df, count, info = fetch_table(table_key)
        if df is None:
            summary_rows.append({
                "表名": table_key,
                "中文标签": table_label,
                "列数": 0,
                "行数": 0,
                "列名列表": "",
                "open_date 字段": "无",
                "open_date 填充数": 0,
                "open_date 填充率": "0%",
            })
            continue
        cols = info["cols"]
        print(f"     列数: {len(cols)}, 行数: {count}")
        print(f"     open_date 填充: {info['open_date_filled']}/{count} ({info['open_date_filled_pct']}%)")
        table_results.append((table_key, table_label, df, count, info))
        summary_rows.append({
            "表名": table_key,
            "中文标签": table_label,
            "列数": len(cols),
            "行数": count,
            "列名列表": ", ".join(cols),
            "open_date 字段": "有" if info["has_open_date"] else "无（schema 中不存在）",
            "open_date 填充数": info["open_date_filled"],
            "open_date 填充率": f"{info['open_date_filled_pct']}%" if info["has_open_date"] else "N/A",
        })

    summary_df = pd.DataFrame(summary_rows)
    total_rows = sum(r["行数"] for r in summary_rows)

    print(f"\n写入 Excel: {out_path}")
    with pd.ExcelWriter(str(out_path), engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="汇总", index=False)
        # openpyxl 不允许 sheet 名含 : \ / ? * [ ]
        for table_key, table_label, df, count, info in table_results:
            sheet_name = f"{table_label}-{table_key}".replace("/", "_").replace("\\", "_")[:31]
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            print(f"  ✓ Sheet: {sheet_name} ({count} 行, {len(info['cols'])} 列)")

    _style_workbook(out_path)

    size_bytes = out_path.stat().st_size
    size_mb = size_bytes / (1024 * 1024)
    print(f"\n✅ 完成")
    print(f"  文件: {out_path}")
    print(f"  大小: {size_mb:.2f} MB ({size_bytes} 字节)")
    print(f"  总行数: {total_rows}")
    print(f"  Sheet 数: {1 + len(table_results)}（汇总 + {len(table_results)} 表）")

    report_path = OUTPUT_DIR / f"export_unified_report_{ts}.txt"
    report_path.write_text(
        f"文件: {out_path}\n"
        f"大小: {size_bytes} bytes ({size_mb:.2f} MB)\n"
        f"总行数: {total_rows}\n"
        f"Sheet 数: {1 + len(table_results)}\n"
        f"导出时间: {datetime.now().isoformat()}\n"
        f"\n--- 各表明细 ---\n" +
        "\n".join(
            f"{r['中文标签']} ({r['表名']}): {r['行数']} 行 × {r['列数']} 列 | "
            f"open_date 字段={r['open_date 字段']} | "
            f"填充数={r['open_date 填充数']} | 填充率={r['open_date 填充率']}"
            for r in summary_rows
        ) + "\n"
    )
    print(f"  报告: {report_path}")
    return str(out_path), total_rows, size_bytes, summary_rows


def _style_workbook(path: Path):
    wb = load_workbook(str(path))
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    alt_fill    = PatternFill("solid", fgColor="EBF3FB")
    thin        = Side(style="thin", color="BFBFBF")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)
    center      = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left        = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    for ws in wb.worksheets:
        for cell in ws[1]:
            cell.fill   = header_fill
            cell.font   = header_font
            cell.alignment = center
            cell.border = border

        max_rows = ws.max_row
        if max_rows > 20000:
            print(f"    ! {ws.title} 行数={max_rows} 较大，跳过隔行底色")
            continue
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=max_rows), start=2):
            fill = alt_fill if row_idx % 2 == 0 else None
            for cell in row:
                if fill:
                    cell.fill = fill
                cell.border = border
                cell.alignment = left
                cell.font = Font(size=9)
        ws.freeze_panes = "A2"
    wb.save(str(path))


if __name__ == "__main__":
    main()