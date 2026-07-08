#!/usr/bin/env python3
"""导出全域招标数据到 Excel（v6）"""
import logging
import sqlite3
from pathlib import Path
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DATA_DIR = Path(__file__).parent / "data"
OUTPUT_DIR = Path(__file__).parent / "output"

SITES = [
    ("jszbcg",       "江苏省政府采购网"),
    ("yancheng_gov", "盐城市政府采购网"),
    ("ycggzy",       "盐城市公共资源交易网"),
    ("bigdata",      "盐城大数据平台"),
    ("jingkai",      "盐城经开区"),
    ("kaifaqu",      "盐城开发区"),
    ("chennan",      "盐南高新区"),
    ("dongfang",     "东方集团"),
    ("dushi",        "都市国际"),
    ("jscn",         "江苏城南"),
    ("yueda",        "悦达集团"),
    ("sufu",         "苏服采"),
]

TYPE_LABEL = {
    "tender":    "招标公告",
    "award":     "中标/成交公告",
    "intention": "采购意向",
    "price_cap": "最高限价公示",
    "other":     "其他",
}

COL_MAP = {
    "std_district":   "标准区县",
    "proj_major_cat": "项目大类",
    "proj_minor_cat": "项目小类",
    "project_name":   "项目名称",
    "notice_type":    "公告类型",
    "region":         "所在区域",
    "purchaser":      "发包单位",
    "budget":         "预算(万元)",      # numeric, /10000
    "budget_text":    "预算情况",        # raw text
    "open_date":      "开标时间",
    "winner":         "中标单位",
    "winning_amount": "中标金额(元)",
    "publish_date":   "发布日期",
    "detail_url":     "详情链接",
}


def load_all() -> pd.DataFrame:
    dfs = []
    for site_key, site_name in SITES:
        db_path = DATA_DIR / f"{site_key}.db"
        if not db_path.exists():
            continue
        db = sqlite3.connect(str(db_path))
        # ycggzy has extra section column
        has_section = site_key == "ycggzy" and any(
            r[1] == "section" for r in db.execute("PRAGMA table_info(notices)").fetchall()
        )
        db_cols = {r[1] for r in db.execute("PRAGMA table_info(notices)").fetchall()}
        cols = [c for c in COL_MAP.keys() if c in db_cols]
        if has_section:
            cols = ["section"] + cols
        # std_district が存在しない DB は NULL 列で補完
        select_exprs = [
            c if c in db_cols else f"NULL AS {c}"
            for c in (["section"] + list(COL_MAP.keys()) if has_section else list(COL_MAP.keys()))
        ]
        rows = db.execute(
            "SELECT " + ",".join(select_exprs) + " FROM notices ORDER BY publish_date DESC"
        ).fetchall()
        cols = (["section"] + list(COL_MAP.keys())) if has_section else list(COL_MAP.keys())
        db.close()
        if not rows:
            continue
        df = pd.DataFrame(rows, columns=cols)
        if "section" not in df.columns:
            df["section"] = ""
        df["site"] = site_name
        dfs.append(df)
    if not dfs:
        return pd.DataFrame()
    all_df = pd.concat(dfs, ignore_index=True)
    all_df["notice_type"] = all_df["notice_type"].map(TYPE_LABEL).fillna("其他")
    # 格式化 open_date
    all_df["open_date"] = all_df["open_date"].apply(
        lambda x: x[:16] if isinstance(x, str) and len(x) >= 16 else x
    )
    # budget: 元 → 万元（2位小数）
    def fmt_budget_wan(v):
        if v is None:
            return ""
        try:
            return round(float(v) / 10000, 2)
        except Exception:
            return ""
    all_df["budget"] = all_df["budget"].apply(fmt_budget_wan)

    # winning_amount: 数值 → 带千位分隔符字符串
    def fmt_amount(v):
        if v is None:
            return ""
        try:
            return f"{float(v):,.2f}"
        except Exception:
            return str(v)
    all_df["winning_amount"] = all_df["winning_amount"].apply(fmt_amount)
    return all_df


def write_excel(df: pd.DataFrame, path: Path):
    OUTPUT_DIR.mkdir(exist_ok=True)

    # 全站汇总 sheet
    summary = []
    for site_key, site_name in SITES:
        db_path = DATA_DIR / f"{site_key}.db"
        if not db_path.exists():
            continue
        db = sqlite3.connect(str(db_path))
        total = db.execute("SELECT COUNT(*) FROM notices").fetchone()[0]
        pc    = db.execute("SELECT COUNT(*) FROM notices WHERE purchaser IS NOT NULL").fetchone()[0]
        bu    = db.execute("SELECT COUNT(*) FROM notices WHERE budget IS NOT NULL").fetchone()[0]
        od    = db.execute("SELECT COUNT(*) FROM notices WHERE open_date IS NOT NULL").fetchone()[0]
        wi    = db.execute("SELECT COUNT(*) FROM notices WHERE winner IS NOT NULL AND winner != ''").fetchone()[0]
        db.close()
        summary.append({
            "站点": site_name, "总条数": total,
            "有发包单位": pc, "有预算": bu, "有开标时间": od, "有中标单位": wi,
        })
    summary_df = pd.DataFrame(summary)

    with pd.ExcelWriter(str(path), engine="openpyxl") as writer:
        # Sheet1: 汇总
        summary_df.to_excel(writer, sheet_name="汇总", index=False)
        data_cols = list(COL_MAP.keys())
        rename = COL_MAP
        # 每个网站一个 sheet
        for site_key, site_name in SITES:
            site_df = df[df["site"] == site_name]
            if site_df.empty:
                continue
            if site_key == "ycggzy" and "section" in site_df.columns:
                out_cols = ["section"] + data_cols
                out_rename = {"section": "版块", **rename}
            else:
                out_cols = data_cols
                out_rename = rename
            sub = site_df[out_cols].rename(columns=out_rename)
            sheet_name = site_name[:31]
            sub.to_excel(writer, sheet_name=sheet_name, index=False)

    # 美化
    _style_workbook(path)
    print(f"✓ 导出: {path}")


def _style_workbook(path: Path):
    wb = load_workbook(str(path))
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    alt_fill    = PatternFill("solid", fgColor="EBF3FB")
    thin        = Side(style="thin", color="BFBFBF")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)
    center      = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left        = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    for ws in wb.worksheets:
        for cell in ws[1]:
            cell.fill   = header_fill
            cell.font   = header_font
            cell.alignment = center
            cell.border = border

        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            fill = alt_fill if row_idx % 2 == 0 else None
            for cell in row:
                if fill:
                    cell.fill = fill
                cell.border = border
                cell.alignment = left
                cell.font = Font(size=9)

        # 自动列宽（粗估）
        for col_cells in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                try:
                    v = str(cell.value) if cell.value else ""
                    # 中文算2，英文算1
                    length = sum(2 if ord(c) > 127 else 1 for c in v)
                    max_len = max(max_len, length)
                except Exception as e:
                    logging.warning(f'[export_excel_sheet] L202 {e}')
            ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

        ws.freeze_panes = "A2"

    wb.save(str(path))


if __name__ == "__main__":
    print("加载数据...")
    df = load_all()
    print(f"总计: {len(df)} 条")
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    out = OUTPUT_DIR / f"盐城市全域招标信息_v6_{ts}.xlsx"
    write_excel(df, out)
    print(f"文件: {out}")
