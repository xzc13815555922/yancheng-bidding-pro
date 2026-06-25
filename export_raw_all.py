#!/usr/bin/env python3
"""导出 12 个原始 db 全列全数据到 Excel（azE 要求看原始 db 情况）

每站 1 个 Sheet + 1 个汇总 Sheet
不做格式化、不做字段映射、不脱敏
"""
import sqlite3
from pathlib import Path
from datetime import datetime
import sys

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DATA_DIR = Path(__file__).parent / "data"
OUTPUT_DIR = Path(__file__).parent / "output"

# 12 个原始站点（不包括 unified.db）
SITES = [
    ("jszbcg",       "江苏省政府采购网"),
    ("yancheng_gov", "盐城市政府采购网"),
    ("ycggzy",       "盐城市公共资源交易网"),
    ("bigdata",      "盐城大数据平台"),
    ("jingkai",      "盐城经开区"),
    ("kaifaqu",      "盐城开发区"),
    ("chennan",      "盐南高新区"),
    ("dongfang",     "东方集团"),
    ("dushi",        "都市国际"),
    ("jscn",         "江苏城南"),
    ("yueda",        "悦达集团"),
    ("sufu",         "苏服采"),
]


def fetch_site_raw(site_key: str, site_name: str):
    """直接从 db SELECT * 拿所有列所有行。"""
    db_path = DATA_DIR / f"{site_key}.db"
    if not db_path.exists():
        return None, None, None
    db = sqlite3.connect(str(db_path))
    try:
        # 拿 schema
        cols_info = db.execute("PRAGMA table_info(notices)").fetchall()
        cols = [r[1] for r in cols_info]
        if not cols:
            db.close()
            return None, None, None
        # SELECT * 所有行
        rows = db.execute(f"SELECT * FROM notices").fetchall()
        count = db.execute("SELECT COUNT(*) FROM notices").fetchone()[0]
        db.close()
        df = pd.DataFrame(rows, columns=cols)
        return df, count, cols
    except Exception as e:
        print(f"  ✗ {site_key} 失败: {e}", file=sys.stderr)
        db.close()
        return None, None, None


def main():
    print("=" * 60)
    print("原始 db 全量导出（azE 任务）")
    print("=" * 60)

    OUTPUT_DIR.mkdir(exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    out_path = OUTPUT_DIR / f"招标公告数据库全量导出_{ts}.xlsx"

    summary_rows = []
    site_results = []  # (site_key, site_name, df, count, cols)

    for site_key, site_name in SITES:
        print(f"  → {site_name} ({site_key}.db)")
        df, count, cols = fetch_site_raw(site_key, site_name)
        if df is None:
            summary_rows.append({
                "站点 key": site_key,
                "站点名称": site_name,
                "db 文件存在": False,
                "列数": 0,
                "行数": 0,
                "列名列表": "",
            })
            continue
        print(f"     列数: {len(cols)}, 行数: {count}")
        site_results.append((site_key, site_name, df, count, cols))
        summary_rows.append({
            "站点 key": site_key,
            "站点名称": site_name,
            "db 文件存在": True,
            "列数": len(cols),
            "行数": count,
            "列名列表": ", ".join(cols),
        })

    summary_df = pd.DataFrame(summary_rows)
    total_rows = sum(r["行数"] for r in summary_rows)

    print(f"\n写入 Excel: {out_path}")
    with pd.ExcelWriter(str(out_path), engine="openpyxl") as writer:
        # Sheet 0: 汇总
        summary_df.to_excel(writer, sheet_name="汇总", index=False)

        # 每站 1 个 Sheet
        for site_key, site_name, df, count, cols in site_results:
            # Sheet 名限 31 字符
            sheet_name = site_name[:31]
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            print(f"  ✓ Sheet: {sheet_name} ({count} 行, {len(cols)} 列)")

    # 美化（轻量，避免巨大文件）
    _style_workbook(out_path)

    # 文件大小
    size_bytes = out_path.stat().st_size
    size_mb = size_bytes / (1024 * 1024)
    print(f"\n✅ 完成")
    print(f"  文件: {out_path}")
    print(f"  大小: {size_mb:.2f} MB ({size_bytes} 字节)")
    print(f"  总行数: {total_rows}")
    print(f"  Sheet 数: {1 + len(site_results)}（汇总 + {len(site_results)} 站）")

    # 写一个 report 文件方便 main session 读
    report_path = OUTPUT_DIR / f"export_raw_report_{ts}.txt"
    report_path.write_text(
        f"文件: {out_path}\n"
        f"大小: {size_bytes} bytes ({size_mb:.2f} MB)\n"
        f"总行数: {total_rows}\n"
        f"Sheet 数: {1 + len(site_results)}\n"
        f"导出时间: {datetime.now().isoformat()}\n"
        f"\n--- 各站明细 ---\n" +
        "\n".join(
            f"{r['站点名称']}: {r['行数']} 行 × {r['列数']} 列 (key={r['站点 key']})"
            for r in summary_rows
        ) + "\n"
    )
    print(f"  报告: {report_path}")

    return str(out_path), total_rows, size_bytes


def _style_workbook(path: Path):
    """轻量样式：冻结首行 + 加粗表头 + 隔行底色，不改列宽（避免巨慢）。"""
    wb = load_workbook(str(path))
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    alt_fill    = PatternFill("solid", fgColor="EBF3FB")
    thin        = Side(style="thin", color="BFBFBF")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)
    center      = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left        = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    for ws in wb.worksheets:
        # 表头
        for cell in ws[1]:
            cell.fill   = header_fill
            cell.font   = header_font
            cell.alignment = center
            cell.border = border

        # 隔行底色 + 边框（不全量遍历 raw_json 大列，限制最大行数避免 OOM）
        max_rows = ws.max_row
        max_cols = ws.max_column
        if max_rows > 20000:
            print(f"    ! {ws.title} 行数={max_rows} 较大，跳过隔行底色")
            continue
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=max_rows), start=2):
            fill = alt_fill if row_idx % 2 == 0 else None
            for cell in row:
                if fill:
                    cell.fill = fill
                cell.border = border
                cell.alignment = left
                cell.font = Font(size=9)

        ws.freeze_panes = "A2"

    wb.save(str(path))


if __name__ == "__main__":
    main()